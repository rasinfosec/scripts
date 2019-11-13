#! python3
import openpyxl # pip3 install openpyxl
from netaddr import * #pip3 install netaddr
import argparse, os

parser = argparse.ArgumentParser()
dir_path = os.path.dirname(os.path.realpath(__file__))


def work(file_in, file_out, ip_column, mask_column, new_column):
    wb = openpyxl.load_workbook(file_in)
    sheet = wb['Sheet1']
    row_count = sheet.max_row
    column_count = sheet.max_column

    for rowNum in range(2, row_count + 1): # skip the first row
        ip = sheet.cell(row=rowNum, column=int(ip_column)).value
        mask = sheet.cell(row=rowNum, column=int(mask_column)).value
        try:
            cidr = IPAddress(mask).netmask_bits()
            change = ('{}/{}'.format(ip, cidr))
            ip_net = IPNetwork(change)
            true_cidr = ip_net.cidr
            sheet.cell(row=rowNum, column=int(new_column)).value = str(true_cidr)
        except:
    	    continue
    wb.save(file_out)


def main():
    parser.add_argument("-f", "--file", dest = "file", help="I need a valid xlsx file.")
    parser.add_argument("-ic", "--ip_column", dest = "ip_column", default = 1, help="What column is the IP located?  Default is column 1.")
    parser.add_argument("-mc", "--mask_column", dest = "mask_column", default = 2, help="What column is the netmask located?  Default is column 2.")
    parser.add_argument("-n", "--new_column", dest = "new_column", default = 3, help="What column should the new data be written?.  Default is column 3.")
    args = parser.parse_args()
    file = args.file
    mask_column = args.mask_column
    ip_column = args.ip_column
    new_column = args.new_column
    file_in = os.path.abspath(dir_path + '/' + file)
    file_out = os.path.abspath(dir_path + '/' + 'changed_' + file)
    work(file_in, file_out, ip_column, mask_column, new_column)
 

if __name__ == '__main__':
    main() 
